"""Run an isolated, virtual-clock cultivation simulation and write JSON/XLSX evidence.

The runner deliberately uses the public V2 population, recruitment, maintenance,
resource, and completion services.  It only owns the clock between service
calls; it does not monkey-patch ``timezone.now``.  This keeps the simulation
close to production behavior while making every timer and settlement explicit.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import traceback
from collections import Counter, defaultdict
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


class SimulationFailure(RuntimeError):
    """Raised after a report was written but one or more invariants failed."""


def _iso(value: datetime | date | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.astimezone(UTC).isoformat()
    return value.isoformat()


def _max_id(model: Any) -> int:
    value = model.objects.order_by("-id").values_list("id", flat=True).first()
    return int(value or 0)


def _counter_payload(counter: Counter[str]) -> dict[str, int]:
    return {str(key): int(value) for key, value in sorted(counter.items())}


def _json_default(value: object) -> str:
    if isinstance(value, (datetime, date)):
        return _iso(value) or ""
    raise TypeError(f"unsupported JSON value: {type(value).__name__}")


def _prepare_isolated_simulation_database() -> None:
    """Prepare the hermetic SQLite catalog required by the public V2 services."""
    from django.conf import settings
    from django.core.management import call_command
    from django.db import connection

    database_engine = str(settings.DATABASES["default"]["ENGINE"])
    if database_engine != "django.db.backends.sqlite3":
        raise SimulationFailure(
            "simulation requires the hermetic SQLite test database; " f"refusing to run against {database_engine!r}"
        )

    if "django_migrations" not in connection.introspection.table_names():
        call_command("migrate", verbosity=0, interactive=False)

    # Migrations intentionally contain only a small bootstrap catalog.  The
    # simulation needs the complete data-driven catalog so candidate and
    # upgrade outcomes are not artifacts of the migration seed.
    call_command("load_troop_templates", verbosity=0, skip_images=True)
    call_command("load_guest_templates", verbosity=0, skip_images=True)
    call_command("load_building_templates", verbosity=0)


class VirtualPlayerSimulation:
    def __init__(
        self,
        *,
        player_count: int,
        days: int,
        seed: int,
        start_at: datetime,
        output_prefix: Path,
    ) -> None:
        if player_count < 1 or player_count > 100:
            raise ValueError("player_count must be between 1 and 100")
        if days < 30:
            raise ValueError("days must be at least 30 so day-15 and day-30 checkpoints are available")
        if start_at.tzinfo is None or start_at.utcoffset() is None:
            raise ValueError("start_at must be timezone-aware")
        self.player_count = int(player_count)
        self.days = int(days)
        self.seed = int(seed)
        self.start_at = start_at.astimezone(UTC)
        self.end_at = self.start_at + timedelta(days=self.days)
        self.output_prefix = output_prefix
        self.profile_ids: tuple[int, ...] = ()
        self.manor_ids: tuple[int, ...] = ()
        self.initial_resources: dict[int, dict[str, int]] = {}
        self.resource_event_baseline_id = 0
        self.schedule_audit: dict[tuple[int, str], list[dict[str, Any]]] = {}
        self.event_log: list[dict[str, Any]] = []
        self.runtime_errors: list[dict[str, Any]] = []
        self.problems: list[dict[str, Any]] = []
        self.checkpoint_snapshots: dict[int, dict[str, Any]] = {}
        self.last_resource_sync_at: datetime | None = None
        self.last_schedule_audit_local_dates: dict[int, date] = {}
        self.recruitment_missing_audit_keys: set[tuple[int, str, int]] = set()
        self.process_count = 0

    def _record_error(self, *, phase: str, now: datetime, exc: BaseException) -> None:
        self.runtime_errors.append(
            {
                "phase": phase,
                "at": _iso(now),
                "type": type(exc).__name__,
                "message": str(exc),
                "traceback": "".join(traceback.format_exception(exc))[-6000:],
            }
        )

    def _configure_runtime(self) -> None:
        from django.conf import settings

        # Keep the population cell bounded to exactly the requested sample.
        # The override is process-local and only affects this isolated test DB.
        settings.VIRTUAL_PLAYER_CONFIG = {
            "population": {
                # Use one explicit search-demand cell so the regional planner
                # materializes exactly this sample instead of distributing a
                # regional floor across every configured region.
                "region_floor": 0,
                "region_active_multiplier": 0,
                "global_floor": self.player_count,
                "global_active_multiplier": 0,
                "hard_cap": self.player_count,
            }
        }
        # Production actions settle their own frozen resource snapshot before
        # spending.  A short batch interval avoids emitting one natural-
        # production event for every unrelated completion while preserving the
        # same balance at every spend and checkpoint boundary.  A six-hour
        # interval is sufficient for the simulation because every consuming
        # domain settles its own frozen production snapshot before spending.
        settings.RESOURCE_SYNC_MIN_INTERVAL_SECONDS = 6 * 60 * 60

        from gameplay.services.virtual_player_core.runtime_preflight import initialize_virtual_player_v2_runtime

        report = initialize_virtual_player_v2_runtime(apply=True)
        failed = [check for check in report.checks if not check.passed]
        if failed:
            raise SimulationFailure(
                "V2 runtime preflight failed: " + "; ".join(f"{check.code}: {check.detail}" for check in failed)
            )

    def _create_population(self) -> None:
        from gameplay.models import BotProfile, Manor
        from gameplay.services.virtual_player_core.backfill import record_virtual_player_backfill_demand
        from gameplay.services.virtual_player_core.config import clear_virtual_player_config_cache
        from gameplay.services.virtual_player_core.population_runtime import (
            PopulationCellReconcileStatus,
            merge_population_recompute_demand,
            reconcile_virtual_player_population_cell,
        )
        from gameplay.services.virtual_player_core.profile_store import reset_virtual_player_simulation_clock

        clear_virtual_player_config_cache()
        self._configure_runtime()
        region = "north"
        prestige_band = "newbie"
        record_virtual_player_backfill_demand(
            region=region,
            prestige_band=prestige_band,
            needed=self.player_count,
        )
        reconcile_results: list[dict[str, Any]] = []
        for _attempt in range(8):
            current_count = BotProfile.objects.filter(engine_version=2, policy_version=2).count()
            if current_count >= self.player_count:
                break
            merge_population_recompute_demand(
                region=region,
                prestige_band=prestige_band,
                now=self.start_at,
            )
            result = reconcile_virtual_player_population_cell(
                region=region,
                prestige_band=prestige_band,
                limit=self.player_count,
                now=self.start_at,
            )
            reconcile_results.append(result.to_payload())
            if result.status not in {
                PopulationCellReconcileStatus.COMPLETED,
                PopulationCellReconcileStatus.CONTINUED,
            }:
                raise SimulationFailure(f"population reconcile failed: {result.to_payload()}")

        profiles = tuple(
            BotProfile.objects.filter(engine_version=2, policy_version=2).select_related("manor").order_by("id")
        )
        if len(profiles) != self.player_count:
            raise SimulationFailure(
                f"population cardinality mismatch after reconcile: expected {self.player_count}, got {len(profiles)}"
            )
        self.profile_ids = tuple(int(profile.id) for profile in profiles)
        self.manor_ids = tuple(int(profile.manor_id) for profile in profiles)

        # Bootstrap rows may carry wall-clock timestamps.  The simulation
        # resets only clock/index fields; the starting assets and resources are
        # left as produced by the real V2 bootstrap path.
        for profile in profiles:
            schedules = self._schedules_for_profile(profile.id, self.start_at)
            next_recruitment_at = schedules[0]["due_at"] if schedules else None
            reset_virtual_player_simulation_clock(
                int(profile.id),
                now=self.start_at,
                next_recruitment_at=next_recruitment_at,
            )
            Manor.objects.filter(pk=profile.manor_id).update(
                resource_updated_at=self.start_at,
                last_active_at=self.start_at,
                action_points_updated_at=self.start_at,
            )

        self.resource_event_baseline_id = _max_id(self._resource_event_model())
        for manor in Manor.objects.filter(pk__in=self.manor_ids).order_by("id"):
            self.initial_resources[int(manor.id)] = {
                "silver": int(manor.silver),
                "grain": int(manor.grain),
            }
        self.population_reconcile_results = reconcile_results

    @staticmethod
    def _resource_event_model() -> Any:
        from gameplay.models import ResourceEvent

        return ResourceEvent

    @staticmethod
    def _schedules_for_profile(profile_id: int, now: datetime) -> list[dict[str, Any]]:
        from gameplay.services.virtual_player_core.recruitment import iter_virtual_recruitment_schedule

        return [
            {
                "quota_date": _iso(schedule.quota_date),
                "quota_ordinal": int(schedule.quota_ordinal),
                "pool_key": str(schedule.pool_key),
                "due_at": schedule.due_at,
                "operation_id": str(schedule.operation_id),
            }
            for schedule in iter_virtual_recruitment_schedule(profile_id, now=now)
        ]

    def _record_schedule_expectations(self, now: datetime) -> None:
        from django.utils import timezone

        from gameplay.models import BotProfile

        local_date = timezone.localtime(now).date()
        for profile_id in self.profile_ids:
            if self.last_schedule_audit_local_dates.get(profile_id) == local_date:
                continue
            profile = BotProfile.objects.filter(pk=profile_id).only("state", "retire_at").first()
            if profile is None:
                continue
            if profile.retire_at is not None and profile.retire_at <= now:
                # Retirement is a valid lifecycle terminal state.  No new
                # daily quota is expected after the profile's retirement
                # instant, so do not manufacture post-retirement omissions.
                self.last_schedule_audit_local_dates[profile_id] = local_date
                continue
            schedules = self._schedules_for_profile(profile_id, now)
            self.last_schedule_audit_local_dates[profile_id] = local_date
            if not schedules:
                continue
            quota_date = str(schedules[0]["quota_date"])
            self.schedule_audit.setdefault(
                (profile_id, quota_date),
                [
                    {
                        **row,
                        "due_at": _iso(row["due_at"]),
                        "observed_at": _iso(now),
                    }
                    for row in schedules
                ],
            )

        # A missing profile is a runtime error rather than a silent omission
        # from the daily-candidate evidence.
        observed_ids = {profile_id for profile_id, _quota_date in self.schedule_audit}
        missing = set(self.profile_ids) - observed_ids
        if missing:
            names = BotProfile.objects.filter(pk__in=missing).values_list("id", "archetype")
            for profile_id, archetype in names:
                self.problems.append(
                    {
                        "severity": "warning",
                        "kind": "recruitment_schedule_missing",
                        "at": _iso(now),
                        "profile_id": int(profile_id),
                        "detail": f"profile {profile_id} ({archetype}) has no daily recruitment schedule",
                    }
                )

    def _sync_resources(self, now: datetime) -> None:
        from gameplay.services.resources import sync_resource_production_batch

        sync_resource_production_batch(self.manor_ids, now=now)
        self.last_resource_sync_at = now

    def _sync_resources_exact(self, now: datetime) -> None:
        from django.conf import settings

        previous_interval = getattr(settings, "RESOURCE_SYNC_MIN_INTERVAL_SECONDS", 0)
        settings.RESOURCE_SYNC_MIN_INTERVAL_SECONDS = 0
        try:
            self._sync_resources(now)
        finally:
            settings.RESOURCE_SYNC_MIN_INTERVAL_SECONDS = previous_interval

    def _finalize_due_domains(self, now: datetime) -> bool:
        from gameplay.models import Building, PlayerTechnology
        from gameplay.services.manor.core import finalize_building_upgrade
        from gameplay.services.technology import finalize_technology_upgrade
        from gameplay.services.virtual_player_core.maintenance_completion import (
            record_virtual_player_maintenance_completion,
        )
        from guests.models import Guest, GuestRecruitment
        from guests.services.training import finalize_guest_training

        finalized_any = False
        due_buildings = list(
            Building.objects.filter(
                manor_id__in=self.manor_ids,
                is_upgrading=True,
                upgrade_complete_at__isnull=False,
                upgrade_complete_at__lte=now,
            )
            .order_by("upgrade_complete_at", "id")
            .select_related("building_type")
        )
        for building in due_buildings:
            origin = building.upgrade_complete_at
            if finalize_building_upgrade(building, now=now, send_notification=False) and origin is not None:
                finalized_any = True
                record_virtual_player_maintenance_completion(
                    manor_id=int(building.manor_id),
                    domain_event_kind="building_upgrade",
                    domain_object_id=int(building.id),
                    origin_completed_at=origin,
                    available_at=now,
                )

        due_technologies = list(
            PlayerTechnology.objects.filter(
                manor_id__in=self.manor_ids,
                is_upgrading=True,
                upgrade_complete_at__isnull=False,
                upgrade_complete_at__lte=now,
            ).order_by("upgrade_complete_at", "id")
        )
        for technology in due_technologies:
            origin = technology.upgrade_complete_at
            if finalize_technology_upgrade(technology, send_notification=False, now=now) and origin is not None:
                finalized_any = True
                record_virtual_player_maintenance_completion(
                    manor_id=int(technology.manor_id),
                    domain_event_kind="technology_upgrade",
                    domain_object_id=int(technology.id),
                    origin_completed_at=origin,
                    available_at=now,
                )

        due_guests = list(
            Guest.objects.filter(
                manor_id__in=self.manor_ids,
                training_complete_at__isnull=False,
                training_complete_at__lte=now,
            ).order_by("training_complete_at", "id")
        )
        for guest in due_guests:
            origin = guest.training_complete_at
            if finalize_guest_training(guest, now=now) and origin is not None:
                finalized_any = True
                record_virtual_player_maintenance_completion(
                    manor_id=int(guest.manor_id),
                    domain_event_kind="guest_training",
                    domain_object_id=int(guest.id),
                    origin_completed_at=origin,
                    available_at=now,
                )

        due_recruitments = list(
            GuestRecruitment.objects.filter(
                manor_id__in=self.manor_ids,
                source=GuestRecruitment.Source.VIRTUAL,
                status=GuestRecruitment.Status.PENDING,
                complete_at__lte=now,
            ).order_by("complete_at", "id")
        )
        from gameplay.services.virtual_player_core.recruitment import finalize_virtual_guest_recruitment

        for recruitment in due_recruitments:
            origin = recruitment.complete_at
            if finalize_virtual_guest_recruitment(int(recruitment.id), now=now) and origin is not None:
                finalized_any = True
                record_virtual_player_maintenance_completion(
                    manor_id=int(recruitment.manor_id),
                    domain_event_kind="guest_recruitment",
                    domain_object_id=int(recruitment.id),
                    origin_completed_at=origin,
                    available_at=now,
                )
        return finalized_any

    def _scan_completions(self, now: datetime) -> None:
        from gameplay.services.virtual_player_core.maintenance_completion import (
            scan_virtual_player_maintenance_completions,
        )

        scan_virtual_player_maintenance_completions(limit=max(100, self.player_count * 32), now=now)

    def _run_recruitment(self, now: datetime) -> None:
        from gameplay.models import BotProfile
        from gameplay.services.virtual_player_core.archetype_pacing import resolve_archetype_pacing
        from gameplay.services.virtual_player_core.config import load_virtual_player_config
        from gameplay.services.virtual_player_core.recruitment import schedule_due_virtual_recruitments
        from gameplay.services.virtual_player_state_policy import VIRTUAL_PROFILE_MAINTAINED_STATES
        from guests.models import GuestRecruitment

        before_id = _max_id(GuestRecruitment)
        try:
            started = schedule_due_virtual_recruitments(
                now=now,
                limit=max(100, self.player_count * 16),
            )
        except Exception as exc:  # Keep the report usable for error analysis.
            self._record_error(phase="recruitment", now=now, exc=exc)
            return
        rows = list(
            GuestRecruitment.objects.filter(
                id__gt=before_id,
                bot_profile_id__in=self.profile_ids,
                source=GuestRecruitment.Source.VIRTUAL,
            ).order_by("id")
        )
        profile_rows: dict[int, list[Any]] = defaultdict(list)
        for row in rows:
            if row.bot_profile_id is not None:
                profile_rows[int(row.bot_profile_id)].append(row)
        config = load_virtual_player_config()
        for profile in BotProfile.objects.filter(pk__in=self.profile_ids).order_by("id"):
            if profile.state not in VIRTUAL_PROFILE_MAINTAINED_STATES:
                continue
            pacing = resolve_archetype_pacing(config, str(profile.archetype))
            rows_for_profile = profile_rows.get(int(profile.id), [])
            if rows_for_profile:
                for row in rows_for_profile:
                    self.event_log.append(
                        {
                            "at": _iso(now),
                            "kind": "recruitment",
                            "profile_id": int(profile.id),
                            "status": str(row.status),
                            "reason": str(row.error_message or "accepted"),
                            "recruitment_id": int(row.id),
                            "quota_date": _iso(row.quota_date),
                            "quota_ordinal": row.quota_ordinal,
                            "result_count": int(row.result_count),
                        }
                    )
            else:
                # The public scan intentionally leaves resource/candidate
                # deferred slots retryable.  Keep that fact visible instead of
                # manufacturing a completed zero-result quota row.
                due = [
                    schedule
                    for schedule in self._schedules_for_profile(int(profile.id), now)
                    if datetime.fromisoformat(str(schedule["due_at"]).replace("Z", "+00:00")) <= now
                ]
                if due:
                    for schedule in due:
                        audit_key = (
                            int(profile.id),
                            str(schedule["quota_date"]),
                            int(schedule["quota_ordinal"]),
                        )
                        if audit_key in self.recruitment_missing_audit_keys:
                            continue
                        self.recruitment_missing_audit_keys.add(audit_key)
                        self.event_log.append(
                            {
                                "at": _iso(now),
                                "kind": "recruitment",
                                "profile_id": int(profile.id),
                                "status": "deferred_without_audit",
                                "reason": "no_recruitment_row_created",
                                "result_count": 0,
                                "started_count": int(started),
                                "quota_date": str(schedule["quota_date"]),
                                "quota_ordinal": int(schedule["quota_ordinal"]),
                                "pacing_schema_version": int(pacing.schema_version),
                            }
                        )
        for row in rows:
            if str(row.status) == GuestRecruitment.Status.COMPLETED:
                origin = row.finished_at or row.complete_at
                if origin is not None:
                    from gameplay.services.virtual_player_core.maintenance_completion import (
                        record_virtual_player_maintenance_completion,
                    )

                    record_virtual_player_maintenance_completion(
                        manor_id=int(row.manor_id),
                        domain_event_kind="guest_recruitment",
                        domain_object_id=int(row.id),
                        origin_completed_at=origin,
                        available_at=now,
                    )
        self._scan_completions(now)

    def _has_due_recruitment(self, now: datetime) -> bool:
        from gameplay.models import BotProfile
        from gameplay.services.virtual_player_state_policy import VIRTUAL_PROFILE_MAINTAINED_STATES

        return BotProfile.objects.filter(
            pk__in=self.profile_ids,
            state__in=VIRTUAL_PROFILE_MAINTAINED_STATES,
            next_recruitment_at__isnull=False,
            next_recruitment_at__lte=now,
        ).exists()

    def _run_maintenance(self, now: datetime) -> None:
        from gameplay.models import BotMaintenanceAttempt
        from gameplay.services.virtual_player_core.maintenance import maintain_due_virtual_players

        before_id = _max_id(BotMaintenanceAttempt)
        try:
            maintain_due_virtual_players(now=now, limit=max(100, self.player_count * 4))
        except Exception as exc:  # A single bad player must remain diagnosable.
            self._record_error(phase="maintenance", now=now, exc=exc)
            return
        attempts = BotMaintenanceAttempt.objects.filter(
            id__gt=before_id,
            profile_id__in=self.profile_ids,
        ).order_by("id")
        for attempt in attempts:
            self.event_log.append(
                {
                    "at": _iso(attempt.started_at),
                    "kind": "maintenance",
                    "profile_id": int(attempt.profile_id),
                    "status": str(attempt.outcome),
                    "action_kind": str(attempt.action_kind or ""),
                    "reason": str(attempt.reason or ""),
                    "attempt_id": int(attempt.id),
                    "cycle_id": str(attempt.cycle_id or ""),
                    "shadow_cost": attempt.shadow_cost,
                }
            )

    def _has_due_maintenance(self, now: datetime) -> bool:
        """Avoid a full maintenance scan for completion-only clock events."""
        from gameplay.models import BotMaintenanceCycle, BotProfile

        open_cycles = {
            int(profile_id): next_slot_due_at
            for profile_id, next_slot_due_at in BotMaintenanceCycle.objects.filter(
                profile_id__in=self.profile_ids,
                status=BotMaintenanceCycle.Status.OPEN,
            ).values_list("profile_id", "next_slot_due_at")
        }
        for profile_id, retire_at, next_growth_at in BotProfile.objects.filter(pk__in=self.profile_ids).values_list(
            "id", "retire_at", "next_growth_at"
        ):
            if retire_at is not None and retire_at <= now:
                return True
            cycle_due_at = open_cycles.get(int(profile_id))
            if cycle_due_at is not None:
                if cycle_due_at <= now:
                    return True
            elif next_growth_at is not None and next_growth_at <= now:
                return True
        return False

    def _process_at(self, now: datetime) -> None:
        from gameplay.services.virtual_player_core.safety_metrics import record_safety_heartbeat

        self.process_count += 1
        if self.process_count > 10000:
            raise SimulationFailure("simulation exceeded 10,000 virtual-clock events")
        record_safety_heartbeat("safety_monitor", now=now)
        from django.conf import settings

        sync_interval = max(0, int(getattr(settings, "RESOURCE_SYNC_MIN_INTERVAL_SECONDS", 0)))
        if self.last_resource_sync_at is None or (now - self.last_resource_sync_at).total_seconds() >= sync_interval:
            self._sync_resources(now)
        if self._finalize_due_domains(now):
            self._scan_completions(now)
        self._record_schedule_expectations(now)
        if self._has_due_recruitment(now):
            self._run_recruitment(now)
        if self._has_due_maintenance(now):
            self._run_maintenance(now)

    def _next_event_at(self, now: datetime) -> datetime | None:
        from gameplay.models import BotMaintenanceCycle, BotProfile, Building, PlayerTechnology
        from guests.models import Guest, GuestRecruitment

        candidates: list[datetime] = []

        def add(value: datetime | None) -> None:
            if value is not None and value > now:
                candidates.append(value)

        for next_growth_at in BotProfile.objects.filter(pk__in=self.profile_ids).values_list(
            "next_growth_at", flat=True
        ):
            add(next_growth_at)
        for next_recruitment_at in BotProfile.objects.filter(pk__in=self.profile_ids).values_list(
            "next_recruitment_at", flat=True
        ):
            add(next_recruitment_at)
        for slot_due_at, decision_at in BotMaintenanceCycle.objects.filter(profile_id__in=self.profile_ids).values_list(
            "next_slot_due_at", "next_decision_at"
        ):
            add(slot_due_at)
            add(decision_at)
        for upgrade_complete_at in Building.objects.filter(
            manor_id__in=self.manor_ids,
            is_upgrading=True,
            upgrade_complete_at__isnull=False,
        ).values_list("upgrade_complete_at", flat=True):
            add(upgrade_complete_at)
        for upgrade_complete_at in PlayerTechnology.objects.filter(
            manor_id__in=self.manor_ids,
            is_upgrading=True,
            upgrade_complete_at__isnull=False,
        ).values_list("upgrade_complete_at", flat=True):
            add(upgrade_complete_at)
        for training_complete_at in Guest.objects.filter(
            manor_id__in=self.manor_ids,
            training_complete_at__isnull=False,
        ).values_list("training_complete_at", flat=True):
            add(training_complete_at)
        for complete_at in GuestRecruitment.objects.filter(
            manor_id__in=self.manor_ids,
            source=GuestRecruitment.Source.VIRTUAL,
            status=GuestRecruitment.Status.PENDING,
            complete_at__isnull=False,
        ).values_list("complete_at", flat=True):
            add(complete_at)
        return min(candidates) if candidates else None

    def _queue_problem_scan(self, checkpoint_at: datetime) -> None:
        from gameplay.models import BotMaintenanceCycle, Building, PlayerTechnology
        from guests.models import Guest, GuestRecruitment

        due_queues = [
            (
                "building_upgrade_due_but_open",
                Building.objects.filter(
                    manor_id__in=self.manor_ids,
                    is_upgrading=True,
                    upgrade_complete_at__isnull=False,
                    upgrade_complete_at__lte=checkpoint_at,
                ),
            ),
            (
                "technology_upgrade_due_but_open",
                PlayerTechnology.objects.filter(
                    manor_id__in=self.manor_ids,
                    is_upgrading=True,
                    upgrade_complete_at__isnull=False,
                    upgrade_complete_at__lte=checkpoint_at,
                ),
            ),
            (
                "guest_training_due_but_open",
                Guest.objects.filter(
                    manor_id__in=self.manor_ids,
                    training_complete_at__isnull=False,
                    training_complete_at__lte=checkpoint_at,
                ),
            ),
            (
                "recruitment_due_but_pending",
                GuestRecruitment.objects.filter(
                    manor_id__in=self.manor_ids,
                    source=GuestRecruitment.Source.VIRTUAL,
                    status=GuestRecruitment.Status.PENDING,
                    complete_at__lte=checkpoint_at,
                ),
            ),
        ]
        for kind, queryset in due_queues:
            for row_id, manor_id in queryset.values_list("id", "manor_id"):
                self.problems.append(
                    {
                        "severity": "error",
                        "kind": kind,
                        "at": _iso(checkpoint_at),
                        "manor_id": int(manor_id),
                        "object_id": int(row_id),
                        "detail": "a due timer remained open after the virtual-clock event pass",
                    }
                )

        # A cycle that repeatedly backs off on a stale precondition while all
        # of its domain completions are already in the past is not a legitimate
        # wait.  Without this check, resource settlement could continue while
        # the report falsely claims that cultivation is progressing.
        open_cycles = (
            BotMaintenanceCycle.objects.filter(
                profile_id__in=self.profile_ids,
                trigger=BotMaintenanceCycle.Trigger.SCHEDULED,
                status=BotMaintenanceCycle.Status.OPEN,
            )
            .select_related("profile")
            .order_by("profile_id", "cycle_ordinal")
        )
        for cycle in open_cycles:
            if cycle.next_slot_due_at is not None and cycle.next_slot_due_at <= checkpoint_at:
                self.problems.append(
                    {
                        "severity": "error",
                        "kind": "maintenance_cycle_due_but_open",
                        "at": _iso(checkpoint_at),
                        "profile_id": int(cycle.profile_id),
                        "manor_id": int(cycle.profile.manor_id),
                        "object_id": int(cycle.id),
                        "detail": "a scheduled maintenance cycle remained due after the virtual-clock event pass",
                    }
                )
            if cycle.last_action_completion_source != "retry_backoff" or cycle.last_reason not in {
                "maintenance_precondition_changed",
                "maintenance_plan_changed",
            }:
                continue
            # A retry whose due time is still in the future is an explicit
            # backoff, not a stalled cycle.  Only report a stale-precondition
            # cycle when the virtual-clock pass has already reached that retry
            # time and it still has no legitimate completion to wait for.
            if cycle.next_slot_due_at is None or cycle.next_slot_due_at > checkpoint_at:
                continue
            payload = cycle.payload if isinstance(cycle.payload, dict) else {}
            retry_history = payload.get("retry_history") or []
            recent_retries = []
            for entry in retry_history:
                if not isinstance(entry, dict):
                    continue
                raw_recorded_at = entry.get("recorded_at")
                if not isinstance(raw_recorded_at, str):
                    continue
                try:
                    recorded_at = datetime.fromisoformat(raw_recorded_at.replace("Z", "+00:00"))
                except ValueError:
                    continue
                if recorded_at <= checkpoint_at:
                    recent_retries.append(entry)
            if len(recent_retries) < 3 or any(
                str(entry.get("reason")) not in {"maintenance_precondition_changed", "maintenance_plan_changed"}
                for entry in recent_retries[-3:]
            ):
                continue
            has_future_pending_completion = False
            for pending_action in payload.get("pending_domain_actions") or []:
                if not isinstance(pending_action, dict):
                    continue
                raw_completion_at = pending_action.get("expected_completion_at")
                if not isinstance(raw_completion_at, str):
                    continue
                try:
                    completion_at = datetime.fromisoformat(raw_completion_at.replace("Z", "+00:00"))
                except ValueError:
                    continue
                if completion_at > checkpoint_at:
                    has_future_pending_completion = True
                    break
            if not has_future_pending_completion:
                self.problems.append(
                    {
                        "severity": "error",
                        "kind": "maintenance_cycle_stalled",
                        "at": _iso(checkpoint_at),
                        "profile_id": int(cycle.profile_id),
                        "manor_id": int(cycle.profile.manor_id),
                        "object_id": int(cycle.id),
                        "detail": "the cycle repeated stale-precondition retries without a future domain completion",
                    }
                )

    def _resource_audit(self, manor_id: int, checkpoint_at: datetime) -> dict[str, Any]:
        from gameplay.models import Manor, ResourceEvent
        from gameplay.services.virtual_player_core.simulation_audit import (
            SimulationAuditError,
            build_resource_ledger_audit,
        )

        final = Manor.objects.get(pk=manor_id)
        events = ResourceEvent.objects.filter(
            manor_id=manor_id,
            id__gt=self.resource_event_baseline_id,
        ).order_by("id")
        try:
            audit = build_resource_ledger_audit(
                initial=self.initial_resources[manor_id],
                final={"silver": int(final.silver), "grain": int(final.grain)},
                events=events,
            )
        except SimulationAuditError as exc:
            self.problems.append(
                {
                    "severity": "error",
                    "kind": "resource_ledger_mismatch",
                    "at": _iso(checkpoint_at),
                    "manor_id": manor_id,
                    "detail": str(exc),
                }
            )
            return {
                "passed": False,
                "error": str(exc),
                "initial": self.initial_resources[manor_id],
                "final": {"silver": int(final.silver), "grain": int(final.grain)},
            }
        return {
            "passed": True,
            "initial": dict(audit.initial),
            "final": dict(audit.final),
            "event_delta": dict(audit.event_delta),
            "by_bucket": {key: dict(value) for key, value in audit.by_bucket.items()},
        }

    def _recruitment_snapshot(self, profile_id: int, checkpoint_at: datetime) -> dict[str, Any]:
        from gameplay.models import BotProfile
        from gameplay.services.virtual_player_core.recruitment import VIRTUAL_RECRUITMENT_RARITY_POLICY_LEGACY_VERSION
        from guests.models import GuestRecruitment

        profile = BotProfile.objects.get(pk=profile_id)
        profile_rows = list(
            GuestRecruitment.objects.filter(
                bot_profile_id=profile_id,
                source=GuestRecruitment.Source.VIRTUAL,
                quota_date__lte=checkpoint_at.date(),
            ).order_by("quota_date", "quota_ordinal", "id")
        )
        expected_until = checkpoint_at
        if profile.retire_at is not None and profile.retire_at < expected_until:
            expected_until = profile.retire_at
        expected_until_iso = _iso(expected_until) or ""
        audited_daily_rows = [
            rows for (row_profile_id, _quota_date), rows in self.schedule_audit.items() if row_profile_id == profile_id
        ]
        strict_due_count = sum(
            1 for rows in audited_daily_rows for row in rows if str(row["due_at"] or "") <= expected_until_iso
        )
        # The production virtual-recruitment contract settles the whole daily
        # batch after the first slot of that day becomes due. Counting only
        # slot-level due_at values would report a deterministic false deficit
        # for the remaining slots in an already-settled batch.
        expected = [
            row
            for rows in audited_daily_rows
            if rows
            and str(sorted(rows, key=lambda item: int(item["quota_ordinal"]))[0]["due_at"] or "") <= expected_until_iso
            for row in rows
        ]
        reason_counts = Counter(str(row.error_message or "accepted") for row in profile_rows)
        rarity_policy_versions: Counter[str] = Counter()
        rarity_policy_snapshots: dict[str, dict[str, Any]] = {}
        rarity_policy_version_by_row: dict[int, int | None] = {}
        for row in profile_rows:
            raw_snapshot = row.pool_snapshot if isinstance(row.pool_snapshot, dict) else {}
            raw_rarity = raw_snapshot.get("rarity")
            if not isinstance(raw_rarity, dict):
                rarity_policy_version_by_row[int(row.id)] = None
                continue
            raw_version = raw_rarity.get("policy_version", VIRTUAL_RECRUITMENT_RARITY_POLICY_LEGACY_VERSION)
            try:
                policy_version = int(raw_version)
            except (TypeError, ValueError):
                policy_version = None
            rarity_policy_version_by_row[int(row.id)] = policy_version
            rarity_policy_versions[str(policy_version) if policy_version is not None else "invalid"] += 1
            policy_payload = {
                "policy_version": policy_version,
                "source": raw_rarity.get("source"),
                "adjustments": raw_rarity.get("adjustments", {}),
                "total_weight": raw_rarity.get("total_weight"),
                "distribution": raw_rarity.get("distribution", []),
            }
            policy_key = json.dumps(policy_payload, ensure_ascii=False, sort_keys=True)
            rarity_policy_snapshots.setdefault(policy_key, policy_payload)
        deferred_count = max(0, len(expected) - len(profile_rows))
        if deferred_count:
            reason_counts["deferred_without_audit"] += deferred_count
        return {
            "expected_daily_slots": len(expected),
            "strict_due_slots": strict_due_count,
            "batch_settled_future_slots": max(0, len(profile_rows) - strict_due_count),
            "recorded_slots": len(profile_rows),
            "deferred_without_audit": deferred_count,
            "expected_until": _iso(expected_until),
            "lifecycle_state": str(profile.state),
            "retire_at": _iso(profile.retire_at),
            "reason_counts": _counter_payload(reason_counts),
            "rarity_policy_versions": _counter_payload(rarity_policy_versions),
            "rarity_policy_snapshots": [rarity_policy_snapshots[key] for key in sorted(rarity_policy_snapshots)],
            "next_recruitment_at": _iso(profile.next_recruitment_at),
            "rows": [
                {
                    "id": int(row.id),
                    "quota_date": _iso(row.quota_date),
                    "quota_ordinal": row.quota_ordinal,
                    "pool_key": str(row.pool.key) if row.pool is not None else None,
                    "rarity_policy_version": rarity_policy_version_by_row.get(int(row.id)),
                    "status": str(row.status),
                    "result_count": int(row.result_count),
                    "error_message": str(row.error_message or ""),
                    "cost": row.cost,
                    "salary_commitment": int(row.salary_commitment),
                    "complete_at": _iso(row.complete_at),
                    "finished_at": _iso(row.finished_at),
                }
                for row in profile_rows
            ],
        }

    def _maintenance_snapshot(self, profile_id: int, checkpoint_at: datetime) -> dict[str, Any]:
        from gameplay.models import BotMaintenanceAttempt

        rows = list(
            BotMaintenanceAttempt.objects.filter(
                profile_id=profile_id,
                started_at__lte=checkpoint_at,
            ).order_by("started_at", "id")
        )
        action_counts = Counter(str(row.action_kind or "no_action") for row in rows)
        reason_counts = Counter(str(row.reason or "applied") for row in rows)
        outcome_counts = Counter(str(row.outcome) for row in rows)
        coverage_gap_counts: Counter[str] = Counter()
        coverage_gap_reason_counts: Counter[str] = Counter()
        coverage_gap_source_counts: Counter[str] = Counter()
        troop_batches: list[dict[str, Any]] = []
        for row in rows:
            shadow_cost = row.shadow_cost if isinstance(row.shadow_cost, dict) else {}
            for raw_gap in shadow_cost.get("coverage_gaps") or ():
                if not isinstance(raw_gap, dict):
                    continue
                action_kind = str(raw_gap.get("action_kind") or "unknown")
                coverage_gap_counts[action_kind] += 1
                coverage_gap_reason_counts[str(raw_gap.get("reason") or "unknown")] += 1
                coverage_gap_source_counts[str(raw_gap.get("reason_source") or "legacy_unknown")] += 1
            if str(row.action_kind) != "troop_recruitment" or str(row.outcome) != "applied":
                continue
            if "troop_batch_quantity" not in shadow_cost:
                continue
            troop_batches.append(
                {
                    "at": _iso(row.started_at),
                    "quantity": int(shadow_cost.get("troop_batch_quantity") or 0),
                    "virtual_grain_cost": int(shadow_cost.get("troop_batch_grain_cost") or 0),
                    "virtual_silver_cost": int(shadow_cost.get("troop_batch_silver_cost") or 0),
                }
            )
        troop_quantities = [int(batch["quantity"]) for batch in troop_batches]
        return {
            "attempt_count": len(rows),
            "execution_count": len(rows),
            "action_counts": _counter_payload(action_counts),
            "reason_counts": _counter_payload(reason_counts),
            "outcome_counts": _counter_payload(outcome_counts),
            "troop_batch_count": len(troop_batches),
            "troop_quantity_total": sum(troop_quantities),
            "troop_quantity_min": min(troop_quantities, default=0),
            "troop_quantity_max": max(troop_quantities, default=0),
            "troop_grain_cost_total": sum(int(batch["virtual_grain_cost"]) for batch in troop_batches),
            "troop_silver_cost_total": sum(int(batch["virtual_silver_cost"]) for batch in troop_batches),
            "troop_batches": troop_batches,
            "coverage_gap_count": sum(coverage_gap_counts.values()),
            "coverage_gap_counts": _counter_payload(coverage_gap_counts),
            "coverage_gap_reason_counts": _counter_payload(coverage_gap_reason_counts),
            "coverage_gap_source_counts": _counter_payload(coverage_gap_source_counts),
            "technology_no_candidate_count": sum(
                1
                for row in rows
                if str(row.action_kind) == "technology_upgrade"
                and str(row.reason)
                in {"no_candidate", "technology_candidate_exhausted", "technology_candidate_blocked"}
            ),
            "technology_insufficient_resource_count": sum(
                1
                for row in rows
                if str(row.action_kind) == "technology_upgrade"
                and str(row.reason) == "technology_insufficient_resource"
            ),
            "technology_capacity_block_count": sum(
                1
                for row in rows
                if str(row.action_kind) == "technology_upgrade"
                and str(row.reason) == "technology_cost_exceeds_capacity"
            ),
        }

    def _player_snapshot(self, profile_id: int, checkpoint_at: datetime) -> dict[str, Any]:
        from gameplay.models import BotProfile, Building, Manor, PlayerTechnology, PlayerTroop, TroopBankStorage
        from gameplay.services.manor.troop_bank import (
            TROOP_BANK_CAPACITY,
            get_troop_bank_remaining_space,
            get_troop_bank_used_space,
        )
        from gameplay.services.manor.troop_capacity import (
            MANOR_TROOP_CAPACITY,
            get_manor_troop_remaining_space,
            get_manor_troop_used_space,
            get_pending_troop_recruitment_space,
        )
        from gameplay.services.virtual_player_core.policy_registry import get_policy_release
        from gameplay.services.virtual_player_core.troop_capacity import (
            get_virtual_troop_remaining_space,
            get_virtual_troop_used_space,
            virtual_troop_capacity_for_prestige_band,
        )
        from guests.models import Guest
        from guests.services.status import resolve_guest_activity_status, resolve_guest_training_state

        profile = BotProfile.objects.select_related("manor").get(pk=profile_id)
        manor = Manor.objects.get(pk=profile.manor_id)
        buildings = list(
            Building.objects.filter(manor_id=manor.id).select_related("building_type").order_by("building_type__key")
        )
        technologies = list(PlayerTechnology.objects.filter(manor_id=manor.id).order_by("tech_key"))
        guests = list(Guest.objects.filter(manor_id=manor.id).select_related("template").order_by("id"))
        troops = list(
            PlayerTroop.objects.filter(manor_id=manor.id)
            .select_related("troop_template")
            .order_by("troop_template__key")
        )
        bank_troops = list(
            TroopBankStorage.objects.filter(manor_id=manor.id)
            .select_related("troop_template")
            .order_by("troop_template__key")
        )
        resource_audit = self._resource_audit(int(manor.id), checkpoint_at)
        recruitment = self._recruitment_snapshot(profile_id, checkpoint_at)
        maintenance = self._maintenance_snapshot(profile_id, checkpoint_at)
        manor_troop_used = get_manor_troop_used_space(manor)
        pending_troop_reservation = get_pending_troop_recruitment_space(manor)
        manor_troop_remaining = get_manor_troop_remaining_space(manor)
        troop_bank_used = get_troop_bank_used_space(manor)
        troop_bank_remaining = get_troop_bank_remaining_space(manor)
        policy_release = get_policy_release(
            version=int(profile.policy_version),
            expected_checksum=str(profile.policy_checksum),
        )
        virtual_troop_capacity = virtual_troop_capacity_for_prestige_band(
            policy_payload=policy_release.payload,
            prestige_band=str(profile.current_prestige_band or profile.prestige_band),
        )
        virtual_troop_used = get_virtual_troop_used_space(manor)
        virtual_troop_remaining = get_virtual_troop_remaining_space(
            manor,
            virtual_capacity=virtual_troop_capacity,
        )
        return {
            "profile_id": int(profile.id),
            "manor_id": int(manor.id),
            "username": str(manor.user.username),
            "archetype": str(profile.archetype),
            "state": str(profile.state),
            "lifecycle": {
                "abandon_at": _iso(profile.abandon_at),
                "retire_at": _iso(profile.retire_at),
            },
            "prestige": int(manor.prestige),
            "prestige_band": str(profile.current_prestige_band or profile.prestige_band),
            "growth_stage": int(profile.growth_stage),
            "next_growth_at": _iso(profile.next_growth_at),
            "buildings": [
                {
                    "key": str(row.building_type.key),
                    "name": str(row.building_type.name),
                    "level": int(row.level),
                    "is_upgrading": bool(row.is_upgrading),
                    "upgrade_complete_at": _iso(row.upgrade_complete_at),
                }
                for row in buildings
            ],
            "technologies": [
                {
                    "key": str(row.tech_key),
                    "level": int(row.level),
                    "is_upgrading": bool(row.is_upgrading),
                    "upgrade_complete_at": _iso(row.upgrade_complete_at),
                }
                for row in technologies
            ],
            "resources": {
                "silver": int(manor.silver),
                "grain": int(manor.grain),
                "silver_capacity": int(manor.silver_capacity),
                "grain_capacity": int(manor.grain_capacity),
                "resource_updated_at": _iso(manor.resource_updated_at),
                "ledger_audit": resource_audit,
            },
            "guests": [
                {
                    "id": int(guest.id),
                    "template_key": str(guest.template.key),
                    "name": str(guest.display_name),
                    "rarity": str(guest.rarity),
                    "level": int(guest.level),
                    "status": resolve_guest_activity_status(guest),
                    "availability_status": str(guest.status),
                    "training_state": resolve_guest_training_state(guest),
                    "force": int(guest.force),
                    "intellect": int(guest.intellect),
                    "defense": int(guest.defense_stat),
                    "agility": int(guest.agility),
                    "training_target_level": int(guest.training_target_level),
                    "training_complete_at": _iso(guest.training_complete_at),
                    "training_remaining_seconds": guest.training_remaining_seconds,
                }
                for guest in guests
            ],
            "guards": {
                # V2 guard actions are committed to PlayerTroop.  Keep the
                # legacy keys as compatibility aliases, but make every
                # report-facing count use this same canonical fact source.
                "troop_count": manor_troop_used,
                "troop_capacity": MANOR_TROOP_CAPACITY,
                "retainer_count": manor_troop_used,
                "retainer_capacity": MANOR_TROOP_CAPACITY,
                "manor_troop_count": manor_troop_used,
                "manor_troop_capacity": MANOR_TROOP_CAPACITY,
                "manor_troop_pending_reservation": pending_troop_reservation,
                "manor_troop_remaining": manor_troop_remaining,
                "troop_bank_count": troop_bank_used,
                "troop_bank_capacity": TROOP_BANK_CAPACITY,
                "troop_bank_remaining": troop_bank_remaining,
                "virtual_troop_count": virtual_troop_used,
                "virtual_troop_capacity": virtual_troop_capacity,
                "virtual_troop_remaining": virtual_troop_remaining,
                "troops": [
                    {
                        "key": str(row.troop_template.key),
                        "name": str(row.troop_template.name),
                        "count": int(row.count),
                        "storage": "manor",
                    }
                    for row in troops
                ],
                "bank_troops": [
                    {
                        "key": str(row.troop_template.key),
                        "name": str(row.troop_template.name),
                        "count": int(row.count),
                        "storage": "bank",
                    }
                    for row in bank_troops
                ],
            },
            "recruitment": recruitment,
            "maintenance": maintenance,
            "queues": {
                "building": [row for row in self._queue_rows("building", manor.id)],
                "technology": [row for row in self._queue_rows("technology", manor.id)],
                "training": [row for row in self._queue_rows("training", manor.id)],
                "recruitment": [row for row in self._queue_rows("recruitment", manor.id)],
            },
        }

    @staticmethod
    def _queue_rows(kind: str, manor_id: int) -> Iterable[dict[str, Any]]:
        from gameplay.models import Building, PlayerTechnology
        from guests.models import Guest, GuestRecruitment

        if kind == "building":
            building_rows = Building.objects.filter(manor_id=manor_id, is_upgrading=True).order_by(
                "upgrade_complete_at", "id"
            )
            return (
                {
                    "id": int(row.id),
                    "complete_at": _iso(row.upgrade_complete_at),
                    "key": str(row.building_type.key),
                }
                for row in building_rows.select_related("building_type")
            )
        if kind == "technology":
            technology_rows = PlayerTechnology.objects.filter(manor_id=manor_id, is_upgrading=True).order_by(
                "upgrade_complete_at", "id"
            )
            return (
                {"id": int(row.id), "complete_at": _iso(row.upgrade_complete_at), "key": str(row.tech_key)}
                for row in technology_rows
            )
        if kind == "training":
            guest_rows = Guest.objects.filter(manor_id=manor_id, training_complete_at__isnull=False).order_by(
                "training_complete_at", "id"
            )
            return ({"id": int(row.id), "complete_at": _iso(row.training_complete_at)} for row in guest_rows)
        recruitment_rows = GuestRecruitment.objects.filter(
            manor_id=manor_id,
            source=GuestRecruitment.Source.VIRTUAL,
            status=GuestRecruitment.Status.PENDING,
        ).order_by("complete_at", "id")
        return ({"id": int(row.id), "complete_at": _iso(row.complete_at)} for row in recruitment_rows)

    def _checkpoint(self, day: int, checkpoint_at: datetime) -> dict[str, Any]:
        from gameplay.models import BotProfile
        from gameplay.services.virtual_player_core.simulation_audit import validate_player_cardinality

        actual_ids = tuple(
            BotProfile.objects.filter(pk__in=self.profile_ids, engine_version=2, policy_version=2)
            .order_by("id")
            .values_list("id", flat=True)
        )
        cardinality = validate_player_cardinality(self.player_count, actual_ids)
        self._queue_problem_scan(checkpoint_at)
        self._sync_resources_exact(checkpoint_at)
        players = [self._player_snapshot(profile_id, checkpoint_at) for profile_id in self.profile_ids]
        aggregates = self._aggregate(players)
        return {
            "day": int(day),
            "at": _iso(checkpoint_at),
            "player_count": cardinality.expected_count,
            "player_ids": list(cardinality.player_ids),
            "players": players,
            "aggregates": aggregates,
        }

    @staticmethod
    def _aggregate(players: list[dict[str, Any]]) -> dict[str, Any]:
        action_counts: Counter[str] = Counter()
        reason_counts: Counter[str] = Counter()
        recruitment_reasons: Counter[str] = Counter()
        total_guests = 0
        total_manor_troops = 0
        total_bank_troops = 0
        manor_capacity_violations = 0
        bank_capacity_violations = 0
        virtual_capacity_violations = 0
        pending_reservation_violations = 0
        troop_counts: Counter[str] = Counter()
        total_troop_batches = 0
        total_troop_quantity = 0
        total_troop_grain_cost = 0
        total_troop_silver_cost = 0
        total_coverage_gap_count = 0
        coverage_gap_counts: Counter[str] = Counter()
        coverage_gap_reason_counts: Counter[str] = Counter()
        coverage_gap_source_counts: Counter[str] = Counter()
        ledger_passed = 0
        tech_levels: Counter[str] = Counter()
        building_levels: Counter[str] = Counter()
        state_counts: Counter[str] = Counter()
        for player in players:
            total_guests += len(player["guests"])
            state_counts[str(player["state"])] += 1
            guards = player["guards"]
            total_manor_troops += int(guards["troop_count"])
            manor_capacity_violations += int(int(guards["manor_troop_count"]) > int(guards["manor_troop_capacity"]))
            bank_capacity_violations += int(int(guards["troop_bank_count"]) > int(guards["troop_bank_capacity"]))
            virtual_capacity_violations += int(
                int(guards["virtual_troop_count"]) > int(guards["virtual_troop_capacity"])
            )
            pending_reservation_violations += int(int(guards["manor_troop_remaining"]) < 0)
            for row in player["guards"]["troops"]:
                troop_counts[str(row["key"])] += int(row["count"])
            for row in player["guards"]["bank_troops"]:
                total_bank_troops += int(row["count"])
            if player["resources"]["ledger_audit"].get("passed"):
                ledger_passed += 1
            action_counts.update(player["maintenance"]["action_counts"])
            reason_counts.update(player["maintenance"]["reason_counts"])
            total_troop_batches += int(player["maintenance"]["troop_batch_count"])
            total_troop_quantity += int(player["maintenance"]["troop_quantity_total"])
            total_troop_grain_cost += int(player["maintenance"]["troop_grain_cost_total"])
            total_troop_silver_cost += int(player["maintenance"]["troop_silver_cost_total"])
            total_coverage_gap_count += int(player["maintenance"].get("coverage_gap_count", 0))
            coverage_gap_counts.update(player["maintenance"].get("coverage_gap_counts", {}))
            coverage_gap_reason_counts.update(player["maintenance"].get("coverage_gap_reason_counts", {}))
            coverage_gap_source_counts.update(player["maintenance"].get("coverage_gap_source_counts", {}))
            recruitment_reasons.update(player["recruitment"]["reason_counts"])
            for row in player["technologies"]:
                tech_levels[row["key"]] += int(row["level"])
            for row in player["buildings"]:
                building_levels[row["key"]] += int(row["level"])
        return {
            "total_guests": total_guests,
            "total_troop_count": total_manor_troops,
            # Compatibility aliases intentionally point to the canonical
            # PlayerTroop total instead of the deprecated Manor counter.
            "total_retainer_count": total_manor_troops,
            "total_manor_troop_count": total_manor_troops,
            "total_bank_troop_count": total_bank_troops,
            "manor_troop_capacity_violations": manor_capacity_violations,
            "troop_bank_capacity_violations": bank_capacity_violations,
            "virtual_troop_capacity_violations": virtual_capacity_violations,
            "pending_troop_reservation_violations": pending_reservation_violations,
            "troop_count_by_key": _counter_payload(troop_counts),
            "troop_batch_count": total_troop_batches,
            "troop_quantity_total": total_troop_quantity,
            "troop_grain_cost_total": total_troop_grain_cost,
            "troop_silver_cost_total": total_troop_silver_cost,
            "coverage_gap_count": total_coverage_gap_count,
            "coverage_gap_counts": _counter_payload(coverage_gap_counts),
            "coverage_gap_reason_counts": _counter_payload(coverage_gap_reason_counts),
            "coverage_gap_source_counts": _counter_payload(coverage_gap_source_counts),
            "state_counts": _counter_payload(state_counts),
            "active_player_count": int(state_counts.get("active", 0)),
            "retired_player_count": int(state_counts.get("retired", 0)),
            "resource_ledger_passed_players": ledger_passed,
            "action_counts": _counter_payload(action_counts),
            "maintenance_reason_counts": _counter_payload(reason_counts),
            "recruitment_reason_counts": _counter_payload(recruitment_reasons),
            "technology_level_sum_by_key": _counter_payload(tech_levels),
            "building_level_sum_by_key": _counter_payload(building_levels),
        }

    def _write_xlsx(self, report: dict[str, Any], path: Path) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        workbook = Workbook()
        workbook.remove(workbook.active)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        body_font = Font(name="Arial", size=10)
        border = Border(bottom=Side(style="thin", color="D9E2F3"))

        def sheet(name: str, headers: list[str], rows: Iterable[list[Any]]) -> None:
            ws = workbook.create_sheet(name)
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in rows:
                ws.append(row)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = body_font
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for column_cells in ws.columns:
                letter = column_cells[0].column_letter
                width = min(48, max(12, max(len(str(cell.value or "")) for cell in column_cells) + 2))
                ws.column_dimensions[letter].width = width

        verification = report["verification"]
        overview_rows = [
            ["simulation_id", report["simulation"]["simulation_id"]],
            ["start_at", report["simulation"]["start_at"]],
            ["end_at", report["simulation"]["end_at"]],
            ["player_count", report["simulation"]["player_count"]],
            ["seed", report["simulation"]["seed"]],
            ["verification_passed", verification["passed"]],
            ["runtime_error_count", len(report["runtime_errors"])],
            ["problem_count", len(report["problems"])],
        ]
        sheet("概览", ["项目", "值"], overview_rows)

        summary_rows: list[list[Any]] = []
        capacity_rows: list[list[Any]] = []
        for checkpoint in report["checkpoints"]:
            for player in checkpoint["players"]:
                guards = player["guards"]
                summary_rows.append(
                    [
                        checkpoint["day"],
                        checkpoint["at"],
                        player["profile_id"],
                        player["username"],
                        player["archetype"],
                        player["state"],
                        player["lifecycle"]["abandon_at"],
                        player["lifecycle"]["retire_at"],
                        player["prestige"],
                        player["growth_stage"],
                        player["resources"]["silver"],
                        player["resources"]["grain"],
                        player["resources"]["silver_capacity"],
                        player["resources"]["grain_capacity"],
                        len(player["guests"]),
                        guards["troop_count"],
                        guards["troop_capacity"],
                        guards["manor_troop_pending_reservation"],
                        guards["manor_troop_remaining"],
                        guards["troop_bank_count"],
                        guards["troop_bank_capacity"],
                        guards["troop_bank_remaining"],
                        guards["virtual_troop_count"],
                        guards["virtual_troop_capacity"],
                        guards["virtual_troop_remaining"],
                        player["maintenance"]["troop_batch_count"],
                        player["maintenance"]["troop_quantity_total"],
                        player["maintenance"]["troop_grain_cost_total"],
                        player["maintenance"]["troop_silver_cost_total"],
                        player["resources"]["ledger_audit"].get("passed"),
                        json.dumps(player["maintenance"]["action_counts"], ensure_ascii=False),
                        json.dumps(player["recruitment"]["reason_counts"], ensure_ascii=False),
                        json.dumps(player["recruitment"]["rarity_policy_versions"], ensure_ascii=False),
                        json.dumps({row["key"]: row["level"] for row in player["buildings"]}, ensure_ascii=False),
                        json.dumps({row["key"]: row["level"] for row in player["technologies"]}, ensure_ascii=False),
                        player["recruitment"]["expected_daily_slots"],
                        player["recruitment"]["recorded_slots"],
                        player["recruitment"]["deferred_without_audit"],
                    ]
                )
                capacity_rows.append(
                    [
                        checkpoint["day"],
                        checkpoint["at"],
                        player["profile_id"],
                        player["username"],
                        guards["manor_troop_count"],
                        guards["manor_troop_capacity"],
                        guards["manor_troop_pending_reservation"],
                        guards["manor_troop_remaining"],
                        guards["troop_bank_count"],
                        guards["troop_bank_capacity"],
                        guards["troop_bank_remaining"],
                        guards["virtual_troop_count"],
                        guards["virtual_troop_capacity"],
                        guards["virtual_troop_remaining"],
                        int(
                            guards["manor_troop_count"] <= guards["manor_troop_capacity"]
                            and guards["troop_bank_count"] <= guards["troop_bank_capacity"]
                            and guards["virtual_troop_count"] <= guards["virtual_troop_capacity"]
                            and guards["manor_troop_remaining"] >= 0
                        ),
                    ]
                )
        sheet(
            "玩家摘要",
            [
                "天数",
                "时间",
                "档案ID",
                "用户名",
                "类型",
                "状态",
                "放弃时间",
                "退役时间",
                "声望",
                "成长阶段",
                "银两",
                "粮食",
                "银库上限",
                "粮仓上限",
                "门客数",
                "护院数",
                "庄园护院上限",
                "庄园护院待完成预留",
                "庄园护院余量",
                "钱庄护院数",
                "钱庄护院上限",
                "钱庄护院余量",
                "声望段护院总数",
                "声望段护院上限",
                "声望段护院余量",
                "护院批次",
                "护院数量累计",
                "护院耗粮累计",
                "护院耗银累计",
                "资源审计通过",
                "培养动作统计",
                "招募原因统计",
                "招募稀有度策略版本",
                "建筑等级",
                "科技等级",
                "应有招募槽",
                "已记录招募槽",
                "未审计槽",
            ],
            summary_rows,
        )
        sheet(
            "护院容量审计",
            [
                "天数",
                "时间",
                "档案ID",
                "用户名",
                "庄园护院数",
                "庄园护院上限",
                "庄园护院待完成预留",
                "庄园护院余量",
                "钱庄护院数",
                "钱庄护院上限",
                "钱庄护院余量",
                "声望段护院总数",
                "声望段护院上限",
                "声望段护院余量",
                "容量校验通过",
            ],
            capacity_rows,
        )

        building_rows: list[list[Any]] = []
        technology_rows: list[list[Any]] = []
        guest_rows: list[list[Any]] = []
        guard_rows: list[list[Any]] = []
        ledger_rows: list[list[Any]] = []
        recruitment_rows: list[list[Any]] = []
        maintenance_rows: list[list[Any]] = []
        for checkpoint in report["checkpoints"]:
            for player in checkpoint["players"]:
                prefix = [checkpoint["day"], checkpoint["at"], player["profile_id"], player["username"]]
                for row in player["buildings"]:
                    building_rows.append(
                        prefix
                        + [row["key"], row["name"], row["level"], row["is_upgrading"], row["upgrade_complete_at"]]
                    )
                for row in player["technologies"]:
                    technology_rows.append(
                        prefix + [row["key"], row["level"], row["is_upgrading"], row["upgrade_complete_at"]]
                    )
                for row in player["guests"]:
                    guest_rows.append(
                        prefix
                        + [
                            row["id"],
                            row["template_key"],
                            row["name"],
                            row["rarity"],
                            row["level"],
                            row["status"],
                            row["availability_status"],
                            row["training_state"],
                            row["force"],
                            row["intellect"],
                            row["defense"],
                            row["agility"],
                            row["training_complete_at"],
                        ]
                    )
                for row in player["guards"]["troops"] + player["guards"]["bank_troops"]:
                    guard_rows.append(prefix + [row["storage"], row["key"], row["name"], row["count"]])
                audit = player["resources"]["ledger_audit"]
                ledger_rows.append(
                    prefix
                    + [
                        audit.get("passed"),
                        json.dumps(audit.get("initial", {}), ensure_ascii=False),
                        json.dumps(audit.get("final", {}), ensure_ascii=False),
                        json.dumps(audit.get("event_delta", {}), ensure_ascii=False),
                        json.dumps(audit.get("by_bucket", {}), ensure_ascii=False),
                        audit.get("error", ""),
                    ]
                )
                for row in player["recruitment"]["rows"]:
                    recruitment_rows.append(
                        prefix
                        + [
                            row["quota_date"],
                            row["quota_ordinal"],
                            row["pool_key"],
                            row["rarity_policy_version"],
                            row["status"],
                            row["result_count"],
                            row["error_message"],
                            json.dumps(row["cost"], ensure_ascii=False),
                            row["salary_commitment"],
                            row["complete_at"],
                            row["finished_at"],
                        ]
                    )
                for row in self.event_log:
                    if (
                        row.get("profile_id") == player["profile_id"]
                        and row.get("kind") == "maintenance"
                        and row.get("at") <= checkpoint["at"]
                    ):
                        maintenance_rows.append(
                            prefix
                            + [
                                row.get("at"),
                                row.get("status"),
                                row.get("action_kind"),
                                row.get("reason"),
                                row.get("target_id"),
                                json.dumps(row.get("shadow_cost", {}), ensure_ascii=False),
                            ]
                        )
        sheet(
            "建筑等级",
            ["天数", "时间", "档案ID", "用户名", "建筑Key", "建筑名", "等级", "升级中", "完成时间"],
            building_rows,
        )
        sheet(
            "科技等级", ["天数", "时间", "档案ID", "用户名", "科技Key", "等级", "升级中", "完成时间"], technology_rows
        )
        sheet(
            "资源流水审计",
            ["天数", "时间", "档案ID", "用户名", "通过", "初始资源", "最终资源", "事件增量", "分类增量", "错误"],
            ledger_rows,
        )
        sheet(
            "门客",
            [
                "天数",
                "时间",
                "档案ID",
                "用户名",
                "门客ID",
                "模板Key",
                "名称",
                "稀有度",
                "等级",
                "状态",
                "可用状态",
                "培养队列状态",
                "武力",
                "智力",
                "防御",
                "敏捷",
                "训练完成时间",
            ],
            guest_rows,
        )
        sheet("护院", ["天数", "时间", "档案ID", "用户名", "存储", "兵种Key", "兵种名", "数量"], guard_rows)
        sheet(
            "招募审计",
            [
                "天数",
                "时间",
                "档案ID",
                "用户名",
                "配额日期",
                "序号",
                "池Key",
                "稀有度策略版本",
                "状态",
                "候选生成数",
                "原因",
                "消耗",
                "工资承诺",
                "完成时间",
                "结束时间",
            ],
            recruitment_rows,
        )
        sheet(
            "维护审计",
            ["天数", "时间", "档案ID", "用户名", "执行时间", "状态", "动作", "原因", "目标ID", "影子成本"],
            maintenance_rows,
        )
        sheet(
            "问题清单",
            ["严重性", "类型", "时间", "档案ID", "庄园ID", "对象ID", "详情"],
            [
                [
                    row.get("severity"),
                    row.get("kind"),
                    row.get("at"),
                    row.get("profile_id"),
                    row.get("manor_id"),
                    row.get("object_id"),
                    row.get("detail"),
                ]
                for row in report["problems"]
            ],
        )
        workbook.save(path)

    def _build_report(self) -> dict[str, Any]:
        checkpoints = [self.checkpoint_snapshots[day] for day in (7, 15, 30)]
        from gameplay.services.virtual_player_core.simulation_audit import validate_player_cardinality

        cardinality = validate_player_cardinality(self.player_count, self.profile_ids)
        verification = {
            "passed": not self.runtime_errors and not any(row.get("severity") == "error" for row in self.problems),
            "player_cardinality": {
                "expected": cardinality.expected_count,
                "actual": len(cardinality.player_ids),
                "player_ids": list(cardinality.player_ids),
            },
            "resource_ledger_passed": all(
                player["resources"]["ledger_audit"].get("passed")
                for checkpoint in checkpoints
                for player in checkpoint["players"]
            ),
            "troop_capacity_passed": all(
                player["guards"]["manor_troop_count"] <= player["guards"]["manor_troop_capacity"]
                and player["guards"]["troop_bank_count"] <= player["guards"]["troop_bank_capacity"]
                and player["guards"]["virtual_troop_count"] <= player["guards"]["virtual_troop_capacity"]
                and player["guards"]["manor_troop_remaining"] >= 0
                for checkpoint in checkpoints
                for player in checkpoint["players"]
            ),
            "troop_report_consistency": all(
                int(player["guards"]["retainer_count"])
                == int(player["guards"]["manor_troop_count"])
                == int(player["guards"]["troop_count"])
                for checkpoint in checkpoints
                for player in checkpoint["players"]
            ),
            "guest_status_consistency": all(
                (
                    player_guest["status"] == "training"
                    if player_guest["training_state"] != "none"
                    else player_guest["status"] == player_guest["availability_status"]
                )
                for checkpoint in checkpoints
                for player in checkpoint["players"]
                for player_guest in player["guests"]
            ),
            "open_due_queue_errors": sum(
                1
                for row in self.problems
                if row.get("kind", "").endswith("_due_but_open")
                or row.get("kind") in {"recruitment_due_but_pending", "maintenance_cycle_stalled"}
            ),
            "maintenance_cycle_stalled_errors": sum(
                1 for row in self.problems if row.get("kind") == "maintenance_cycle_stalled"
            ),
        }
        verification["passed"] = bool(
            verification["passed"]
            and verification["resource_ledger_passed"]
            and verification["troop_capacity_passed"]
            and verification["troop_report_consistency"]
            and verification["guest_status_consistency"]
        )
        return {
            "simulation": {
                "simulation_id": f"vp-sim-{self.start_at.strftime('%Y%m%dT%H%M%SZ')}-{self.seed}",
                "start_at": _iso(self.start_at),
                "end_at": _iso(self.end_at),
                "player_count": self.player_count,
                "days": self.days,
                "seed": self.seed,
                "profile_ids": list(self.profile_ids),
                "manor_ids": list(self.manor_ids),
                "population_reconcile": self.population_reconcile_results,
                "process_count": self.process_count,
            },
            "verification": verification,
            "checkpoints": checkpoints,
            "schedule_audit": [
                {
                    "profile_id": profile_id,
                    "quota_date": quota_date,
                    "expected_slots": rows,
                }
                for (profile_id, quota_date), rows in sorted(self.schedule_audit.items())
            ],
            "event_log": self.event_log,
            "runtime_errors": self.runtime_errors,
            "problems": self.problems,
        }

    def run(self) -> dict[str, Any]:
        self._create_population()
        current = self.start_at
        self._process_at(current)
        checkpoint_times = {self.start_at + timedelta(days=day): day for day in (7, 15, 30) if day <= self.days}
        while current < self.end_at:
            next_at = self._next_event_at(current)
            if next_at is None or next_at > self.end_at:
                next_at = self.end_at
            pending_checkpoints = [checkpoint_at for checkpoint_at in checkpoint_times if checkpoint_at > current]
            if pending_checkpoints:
                next_at = min(next_at, min(pending_checkpoints))
            if next_at <= current:
                raise SimulationFailure(f"virtual clock did not advance after {current.isoformat()}")
            current = next_at
            self._process_at(current)
            checkpoint_day = checkpoint_times.get(current)
            if checkpoint_day is not None:
                self.checkpoint_snapshots[checkpoint_day] = self._checkpoint(checkpoint_day, current)
        if self.process_count <= 0 or set(self.checkpoint_snapshots) != {7, 15, 30}:
            raise SimulationFailure("simulation did not process checkpoints")
        report = self._build_report()
        self.output_prefix.parent.mkdir(parents=True, exist_ok=True)
        json_path = self.output_prefix.with_suffix(".json")
        summary_path = self.output_prefix.with_name(self.output_prefix.name + "_summary.json")
        xlsx_path = self.output_prefix.with_suffix(".xlsx")
        json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=_json_default), encoding="utf-8")
        summary = {
            "simulation": report["simulation"],
            "verification": report["verification"],
            "checkpoint_summaries": [
                {"day": row["day"], "at": row["at"], "aggregates": row["aggregates"]} for row in report["checkpoints"]
            ],
            "runtime_errors": report["runtime_errors"],
            "problems": report["problems"],
        }
        summary_path.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2, default=_json_default), encoding="utf-8"
        )
        self._write_xlsx(report, xlsx_path)
        print(
            json.dumps(
                {
                    "json": str(json_path),
                    "xlsx": str(xlsx_path),
                    "summary": str(summary_path),
                    "passed": report["verification"]["passed"],
                },
                ensure_ascii=False,
            )
        )
        if not report["verification"]["passed"]:
            raise SimulationFailure("simulation completed with failed verification; inspect the generated report")
        return report


def _parse_start_at(raw: str | None) -> datetime:
    if raw:
        value = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    else:
        from django.utils import timezone

        value = timezone.now()
    if value.tzinfo is None or value.utcoffset() is None:
        raise ValueError("--start-at must be timezone-aware ISO-8601")
    return value.astimezone(UTC).replace(second=0, microsecond=0)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Run an isolated V2 virtual-player cultivation simulation")
    parser.add_argument("--players", type=int, default=10)
    parser.add_argument("--days", type=int, default=30)
    parser.add_argument("--seed", type=int, default=20260816)
    parser.add_argument("--start-at", default=None, help="timezone-aware ISO-8601 virtual start time")
    parser.add_argument(
        "--output-prefix",
        default=str(PROJECT_ROOT / "virtual_player_simulation_15d_30d_2026-08-16"),
    )
    args = parser.parse_args(argv)
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
    # ``config.settings.testing`` is an override module, not a standalone
    # settings package.  Mark this process as a test-like run before Django
    # imports ``config.settings`` so it selects SQLite/LocMem/Celery-memory
    # infrastructure without touching a real service.
    simulation_marker = "test"
    if "test" not in sys.argv:
        sys.argv.append(simulation_marker)
    import django

    django.setup()
    logging.getLogger("gameplay.services.virtual_player_core.maintenance").setLevel(logging.WARNING)
    if simulation_marker in sys.argv:
        sys.argv.remove(simulation_marker)
    try:
        _prepare_isolated_simulation_database()
        simulation = VirtualPlayerSimulation(
            player_count=args.players,
            days=args.days,
            seed=args.seed,
            start_at=_parse_start_at(args.start_at),
            output_prefix=Path(args.output_prefix).resolve(),
        )
        simulation.run()
    except Exception as exc:
        print(f"simulation failed: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
